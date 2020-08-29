from flask import Flask, render_template, jsonify, request
from selenium import webdriver
from datetime import date, timedelta
import pyautogui
import time
from pymongo import MongoClient
import openpyxl as xl
app = Flask(__name__)

client = MongoClient('localhost', 27017)
db = client.dbhomework

driver = webdriver.Chrome('C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe')

driver.get('https://mycompany.hyundaicard.com')

pyautogui.PAUSE = 1
pyautogui.FAILSAFE = True

# 팝업 없애기
time.sleep(1)
driver.find_element_by_xpath('//*[@id="layer_type_01"]/div[2]/button').click()

# 팝업 없애기
time.sleep(1)
driver.find_element_by_xpath('//*[@id="layer_type_02"]/div[2]/button').click()

#pyautogui.moveTo(1173,350,1)
#pyautogui.doubleClick(x=1173,y=350, button='left')

#현대카드 홈페이지 내에 메뉴 화면 없애기
pyautogui.moveTo(33,844,1)
#pyautogui.click(x=33,y=844,button='left')

#아이디 화면 띄우기
time.sleep(1)
driver.find_element_by_xpath('//*[@id="contents"]/div[2]/div/div[1]/div/fieldset/div/div[2]/div/span[1]/label/em').click()

#아이디 화면 띄우기
#pyautogui.moveTo(153,564,1)
#pyautogui.doubleClick(x=153,y=564, button='left')

#아이디 입력하기
time.sleep(1)
driver.find_element_by_id("corp_web_mbr_id").send_keys('kuksan88')

time.sleep(10)

# My account
time.sleep(1)
driver.find_element_by_xpath('//*[@id="gnb"]/div/div/ul/li[1]/h3/a').click()
# 카드이용 내역
time.sleep(1)
driver.find_element_by_xpath('//*[@id="gnb"]/div/div/ul/li[1]/div/div/div[1]/div/ul/li[2]/a').click()

# 직접입력
time.sleep(1)
driver.find_element_by_xpath('//*[@id="srchForm"]/fieldset/div[1]/div[5]/div/div/div/div/ul/li[4]/a').click()

# 매월, 매초 구하기
today = date.today()
first_day = today.replace(day=1)

last_day_month_ago = first_day - timedelta(days=1)
first_day_month_ago = last_day_month_ago.replace(day=1)

time.sleep(1)
driver.find_element_by_id("strtDt").click()
time.sleep(1)
pyautogui.write(first_day_month_ago.strftime('%Y%m%d'), interval=0.5)

time.sleep(1)
driver.find_element_by_id("endDt").click()
time.sleep(1)
pyautogui.write(last_day_month_ago.strftime('%Y%m%d'), interval=0.5)

# 확인 버튼
time.sleep(1)
driver.find_element_by_xpath('//*[@id="formCal"]/button').click()

# 조회 버튼
time.sleep(1)
driver.find_element_by_xpath('//*[@id="srchForm"]/fieldset/div[2]/button').click()

# 조회 전체 버튼
time.sleep(1)
driver.find_element_by_xpath('//*[@id="gridbox"]/div[1]/table/tbody/tr[2]/td[1]/div/label').click()

# 매출전표 다운로드 버튼
time.sleep(1)
driver.find_element_by_xpath('//*[@id="approve_article"]/div[4]/div[3]/div/span/a[1]').click()

# 매출전표 다운로드 실시
time.sleep(1)
driver.find_element_by_xpath('//*[@id="layout_popupDiv"]/div/div[2]/div/div/div/div/a[2]').click()

# 엑셀다운로드 다운로드 실시
time.sleep(20)
driver.find_element_by_xpath('//*[@id="approve_article"]/div[4]/div[3]/div/span/a[3]').click()

# 현대카드 내리기
#pyautogui.click(x=1175,y=37, button='left', duration=1)

path1 = "C:/test/"+"copy_"+today.strftime('%Y%m%d')+".xlsx"
path2 = "C:/test/paste.xlsx"

wb1 = xl.load_workbook(filename=path1)
ws1 = wb1.active

wb2 = xl.load_workbook(filename=path2)
# ws = wb.get_sheet_by_name("Sheet1")
ws2 = wb2.active

# copying the cell values from source
# excel file to destination excel file
for i in range(2, 28):
    for j in range(4, 5):
        # reading cell value from source excel file
        c = ws1.cell(row=i, column=j)
        ws2.cell(row=i, column=j).value = c.value
        wb2.save(path2)
